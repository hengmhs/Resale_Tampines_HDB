import geopy.distance
from openpyxl import load_workbook, Workbook, utils

# tampines_east_mrt = (1.3558290924997582, 103.95520090009265)
# tampines_west_mrt = (1.345808393511592, 103.9382382867281)

wb_name = 'resale2013_2018.xlsx'

#simei_mrt = (1.343097244672535, 103.95331048465245)
upper_changi_mrt = (1.34164302491625, 103.96146762682966)

# Load the Workbook
wb = load_workbook(wb_name)
# Load the Worksheet
sheet = wb['Sheet1']

coord_list = []
simei_dist_list = []
changi_dist_list = []

# start from 2 to avoid the column headers, range() is one less step than max
# get coordinates of each place
for i in range(2, sheet.max_row + 1):
    lat = sheet.cell(row=i, column=11).value
    longitude = sheet.cell(row=i, column=12).value
    coord = (lat, longitude)
    coord_list.append(coord)

'''
# get distances between coord and tampines east mrt
for index, coord in enumerate(coord_list):
    print(str(index+1)+'/'+str(len(coord_list)))
    dist = geopy.distance.geodesic(coord, tampines_east_mrt).km
    east_dist_list.append(dist)

# write tampines east mrt to excel
for i in range(0, len(east_dist_list)):
    cell = sheet.cell(column=14, row=i+2)
    cell.value = east_dist_list[i]
'''

# get distances between coord and simei mrt
for index, coord in enumerate(coord_list):
    print(str(index+1)+'/'+str(len(coord_list)),end='\r')
    dist = geopy.distance.geodesic(coord, upper_changi_mrt).km
    changi_dist_list.append(dist)

# write simei mrt to excel
for i in range(0, len(changi_dist_list)):
    cell = sheet.cell(column=utils.column_index_from_string('Q'), row=i+2)
    cell.value = changi_dist_list[i]

wb.save(wb_name)