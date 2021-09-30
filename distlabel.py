from geopy.geocoders import GoogleV3
import geopy.distance
from openpyxl import load_workbook, Workbook

with open('API_KEY.txt','r') as API:
    API_KEY = API.read()

# need to enable Geocoding API in Google Cloud Platform first
geolocator = GoogleV3(API_KEY)

# Tampines MRT lat: 1.3533834 long: 103.9451538
tampines_mrt = (1.3533834, 103.9451538)

wb_name = 'resale2013_2018.xlsx'

# Load the Workbook
wb = load_workbook(wb_name)
# Load the Worksheet
sheet = wb['Sheet1']

addr_list = []
dist_list = []

# start from 2 to avoid the column headers, range() is one less step than max
for i in range(2, sheet.max_row + 1):
    block = str(sheet.cell(row=i, column=4).value)
    street_name = sheet.cell(row=i, column=5).value
    addr_list.append(block + ' ' + street_name)

addr_list_len = len(addr_list)

# list of coordinates
coord_list = []

# save repeat addresses to save API calls
coord_obj = {}

# saves repeat addresses to save API calls
dist_obj = {}

for index, address in enumerate(addr_list):
    print(str(index+1)+'/'+str(addr_list_len))
    if address in dist_obj:
        # get distance
        dist_list.append(dist_obj[address])
        # get coordinates
        coord_list.append(coord_obj[address])
        print('duplicate')
    else:
        location = geolocator.geocode(address)
        place = (location.latitude, location.longitude)
        coord_obj[address] = place
        # get distance
        dist = geopy.distance.geodesic(place, tampines_mrt).km
        # save distance to dist_obj
        dist_obj[address] = dist
        dist_list.append(dist)
        # add coordinates to list 
        coord_list.append(place)


dist_list_len = len(dist_list)
coord_list_len = len(coord_list)

# write distances to excel
# Excel row starts at 2 but index of dist_list must start at 0
# therefore add 2 to row value
for i in range(0,dist_list_len):
    cell = sheet.cell(column=13, row=i+2)
    cell.value = dist_list[i]

# write latitude coordinates to excel
for i in range(0,coord_list_len):
    cell = sheet.cell(column=11, row=i+2)
    cell.value = coord_list[i][0]

# write longitude coordinates to excel
for i in range(0,coord_list_len):
    cell = sheet.cell(column=12, row=i+2)
    cell.value = coord_list[i][1]


wb.save(wb_name)