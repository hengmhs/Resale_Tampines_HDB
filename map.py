import gmplot
from openpyxl import load_workbook, Workbook, utils

wb_name = 'resale2013_2018.xlsx'

# Load the Workbook
wb = load_workbook(wb_name)
# Load the Worksheet
sheet = wb['Sheet1']

with open('API_KEY.txt','r') as API:
    API_KEY = API.read()

gmap = gmplot.GoogleMapPlotter(1.3533834, 103.9451538, 15, apikey=API_KEY)


coord_list = [] # list of all hdbs not near MRT
tamp_list = [] # list of hdbs near tampines MRT
other_list = [] # list of hdbs near another MRT that is not tampines

# start from 2 to avoid the column headers, range() is one less step than max
# get coordinates of each place
for i in range(2, sheet.max_row + 1):
    lat = sheet.cell(row=i, column=utils.column_index_from_string('K')).value
    longitude = sheet.cell(row=i, column=utils.column_index_from_string('L')).value
    coord = (lat, longitude)
    # get value of tampines_mrt, east_mrt, west_mrt
    tampines_mrt = sheet.cell(row=i,column=utils.column_index_from_string('R')).value
    east_mrt = sheet.cell(row=i,column=utils.column_index_from_string('S')).value
    west_mrt = sheet.cell(row=i,column=utils.column_index_from_string('T')).value
    simei_mrt = sheet.cell(row=i,column=utils.column_index_from_string('U')).value
    upper_changi_mrt = sheet.cell(row=i,column=utils.column_index_from_string('V')).value
    # check if tampines_mrt = 1
    if tampines_mrt:
        tamp_list.append(coord)
    # check if coord is near other mrts
    elif east_mrt or west_mrt or simei_mrt or upper_changi_mrt:
        other_list.append(coord)
    else:
        coord_list.append(coord)

# removes duplicates
tamp_set = set(tamp_list)
# convert set back to list
tamp_list = list(tamp_set) 

tamp_list_lat = []
tamp_list_long = []

for coord in tamp_list:
    tamp_list_lat.append(coord[0])
    tamp_list_long.append(coord[1])

# removes duplicates
coord_set = set(coord_list)
# convert set back to list
coord_list = list(coord_set)

coord_list_lat = []
coord_list_long = []

other_set = set(other_list) # removes duplicates
other_list = list(other_set) # convert set back to list

other_list_lat = []
other_list_long = []

for coord in other_list:
    other_list_lat.append(coord[0])
    other_list_long.append(coord[1])

for coord in coord_list:
    coord_list_lat.append(coord[0])
    coord_list_long.append(coord[1])

# scatter takes a list of lats and longs
gmap.scatter(tamp_list_lat,tamp_list_long,size=8,color='red',marker=False,alpha=1)
gmap.scatter(coord_list_lat, coord_list_long,size=8,color='blue',marker=False)
gmap.scatter(other_list_lat, other_list_long,size=8,color='green',marker=False)

# draw Tampines MRT Circle
gmap.circle(1.3533834, 103.9451538,400,face_alpha=0,edge_color='red')

# Tampines West
gmap.circle(1.3558290924997582, 103.95520090009265,400,face_alpha=0,edge_color='green')
# Tampines East
gmap.circle(1.345808393511592, 103.9382382867281,400,face_alpha=0,edge_color='green')
# Simei
gmap.circle(1.343097244672535, 103.95331048465245,400,face_alpha=0,edge_color='green')
# Upper Changi
gmap.circle(1.34164302491625, 103.96146762682966,400,face_alpha=0,edge_color='green')

# output the visualisation
gmap.draw('visualisation.html')